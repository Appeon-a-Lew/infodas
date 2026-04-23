"""
Asset-Based GS → GS++ Gap Analysis
=====================================
Logic:
  1. Load asset GS requirements from gs_check.csv
  2. Match asset name to GS++ Zielobjekt category via embedding
  3. Look up GS++ controls via final_mapping.csv (gs_id based)
  4. Classify pass/fail using Umsetzungsgrad + modalverb + level → priority score

Priority scoring:
  1  Non-Complied + erhöht
  2  Non-Complied + normal-SdT
  3  Optional + erhöht
  4  Optional + normal-SdT
  5  Complied + erhöht
  6  Complied + normal-SdT

Usage:
  python3 src/asset_gap.py --asset "Virtualisierungsserver"
  python3 src/asset_gap.py --all
  python3 src/asset_gap.py --all --json-file mapping.json
"""

import re
import os
import json
import pickle
import argparse
import numpy as np
import pandas as pd
from pathlib import Path
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# ─── PATHS ───────────────────────────────────────────────────────────────────
BASE          = Path("/Users/tarikosmanmutlu/IdeaProjects/Infodas2")
CSV           = BASE / "csv"
NS            = CSV / "namespaces"
JSON_DIR      = BASE / "json"
OUT           = BASE / "output"
OUT.mkdir(parents=True, exist_ok=True)

GS_CHECK_CSV   = CSV / "gs_check.csv"
FINAL_MAPPING  = CSV / "final_mapping.csv"
ZIELOBJEKT_CSV = NS  / "target_object_categories.csv"
GSPP_JSON      = JSON_DIR / "Grundschutz++-catalog.json"
CACHE_FILE     = BASE / "embeddings_cache_asset_gap.pkl"

EMBED_MODEL    = "text-embedding-3-large"
GENERAL_ASSETS = {"Informationsverbund"}
UZ_PRIORITY    = {"Nein": 0, "Teilweise": 1, "Ja": 2, "Entbehrlich": 3}

# ─── STATUS + PRIORITY ───────────────────────────────────────────────────────
STATUS_STYLES = {
    "🔴 Non-Complied": ("FFCCCC", "C00000"),
    "✅ Complied":     ("E2EFDA", "375623"),
    "🔵 Optional":     ("DAE8FC", "1F4E79"),
    "⚠️ Exempted":    ("FFF2CC", "7F6000"),
}

LEGEND_ROWS = [
    ("🔴 Non-Complied", "Not implemented — mandatory",   "MUSS",   "Nein or Teilweise", "Immediate action — P1/P2"),
    ("🔴 Non-Complied", "Not implemented — mandatory",   "SOLLTE", "Nein or Teilweise", "Action required — P3/P4"),
    ("⚠️ Exempted",    "Exempted in GS — review needed","Any",    "Entbehrlich",        "Justify exemption — P7"),
    ("🔵 Optional",     "Optional control",              "KANN",   "Any",                "Risk-based decision — P5/P6"),
    ("✅ Complied",     "Control satisfied",             "Any",    "Ja",                 "No action required"),
]

def compute_priority(status: str, modalverb: str, level: str) -> int:
    """
    Priority 1 = most critical, 7 = least urgent.
    Complied controls get no priority (None).
    """
    if status == "✅ Complied":
        return None
    if status == "⚠️ Exempted":
        return 7
    level_rank = 0 if level == "erhöht" else 1
    base = {
        ("🔴 Non-Complied", "MUSS"):   0,
        ("🔴 Non-Complied", "SOLLTE"): 2,
        ("🔵 Optional",     "KANN"):   4,
    }
    return base.get((status, modalverb), 4) + level_rank + 1


# ─── EMBEDDING HELPERS ───────────────────────────────────────────────────────

def load_cache() -> dict:
    if CACHE_FILE.exists():
        with open(str(CACHE_FILE), "rb") as f:
            return pickle.load(f)
    return {}

def save_cache(cache: dict):
    with open(str(CACHE_FILE), "wb") as f:
        pickle.dump(cache, f)

def get_embeddings(texts: list, cache: dict) -> np.ndarray:
    from openai import OpenAI
    client  = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    missing = [t for t in texts if t not in cache]
    if missing:
        for i in range(0, len(missing), 100):
            batch = missing[i:i+100]
            resp  = client.embeddings.create(input=batch, model=EMBED_MODEL)
            for text, obj in zip(batch, resp.data):
                cache[text] = obj.embedding
        save_cache(cache)
    return np.array([cache[t] for t in texts])


# ─── ZIELOBJEKT MATCHING (embedding) ─────────────────────────────────────────

def match_zielobjekt(asset_name: str, zo_df: pd.DataFrame, cache: dict) -> list[dict]:
    """
    Match asset name to GS++ Zielobjekt categories via embedding similarity.
    Returns top matches sorted by score.
    """
    # Build embed text for each Zielobjekt: name + definition + synonyms
    zo_texts = []
    for _, row in zo_df.iterrows():
        text = row["Zielobjekt"]
        if row.get("Definition"): text += ". " + row["Definition"]
        if row.get("Synonyme"):   text += ". " + row["Synonyme"]
        zo_texts.append(text.strip())

    # Asset embed: just the name (strip code prefix like "S007 ")
    parts       = asset_name.split(None, 1)
    asset_clean = parts[1] if len(parts) > 1 and re.match(r'^[a-zA-Z0-9]+$', parts[0]) else asset_name

    all_texts    = [asset_clean] + zo_texts
    embeddings   = get_embeddings(all_texts, cache)
    asset_emb    = embeddings[0:1]
    zo_embs      = embeddings[1:]

    sims = cosine_similarity(asset_emb, zo_embs)[0]

    results = []
    for i, (_, row) in enumerate(zo_df.iterrows()):
        results.append({
            "zielobjekt": row["Zielobjekt"],
            "typ":        row.get("Typ",""),
            "kategorie":  row.get("Kategorie",""),
            "score":      round(float(sims[i]), 3),
        })
    return sorted(results, key=lambda x: x["score"], reverse=True)


# ─── MODALVERB FROM CATALOG ───────────────────────────────────────────────────

def load_modalverb_map() -> dict:
    """Extract {gspp_id: modalverb} from GS++ catalog JSON."""
    def get_mv(text):
        m = re.search(r'\b(MUSS|SOLLTE|KANN)\b', text or "")
        return m.group(1) if m else "UNKNOWN"

    def get_stmt(ctrl):
        for p in ctrl.get("parts", []):
            if p.get("name") == "statement":
                return p.get("prose", "")
        return ""

    with open(str(GSPP_JSON), encoding="utf-8") as f:
        catalog = json.load(f)

    mv_map = {}
    for g in catalog["catalog"]["groups"]:
        for sg in g.get("groups", []):
            for c in sg.get("controls", []):
                mv_map[c["id"]] = get_mv(get_stmt(c))
                for sc in c.get("controls", []):
                    mv_map[sc["id"]] = get_mv(get_stmt(sc))
    return mv_map


# ─── LOAD FINAL MAPPING ───────────────────────────────────────────────────────

def load_final_mapping(path: Path = FINAL_MAPPING, mv_map: dict = None) -> pd.DataFrame:
    df = pd.read_csv(str(path), dtype=str).fillna("")
    df["gs_ids_list"]  = df["gs_ids"].apply(
        lambda x: [g.strip() for g in x.split(";") if g.strip()]
    )
    df["gspp_practice"] = df["gspp_id"].str.split(".").str[0]

    # Join modalverb from catalog
    if mv_map:
        df["modalverb"] = df["gspp_id"].map(mv_map).fillna("UNKNOWN")
    else:
        df["modalverb"] = "UNKNOWN"

    print(f"  Mapping: {len(df)} GS++ controls | modalverb: {df['modalverb'].value_counts().to_dict()}")
    return df


# ─── GS CHECK ────────────────────────────────────────────────────────────────

def load_gs_check_csv() -> pd.DataFrame:
    return pd.read_csv(str(GS_CHECK_CSV), sep=";", encoding="utf-8-sig", dtype=str).fillna("")

def get_asset_requirements(asset_name: str, gs_check: pd.DataFrame) -> pd.DataFrame:
    asset_lower = asset_name.lower().strip()
    parts       = asset_lower.split(None, 1)
    asset_clean = parts[1] if len(parts) > 1 and re.match(r'^[a-z0-9]+$', parts[0]) else asset_lower
    mask = (
            gs_check["Zielobjekt"].str.lower().str.contains(asset_clean, regex=False, na=False) |
            (gs_check["Zielobjekt"].str.lower() == asset_lower)
    )
    return gs_check[mask].copy()

def aggregate_requirements(gs_asset: pd.DataFrame) -> pd.DataFrame:
    gs_asset = gs_asset.copy()
    gs_asset["_anf_id"] = gs_asset["Anforderung"].str.strip().str.split().str[0]
    rows = []
    for anf_id, grp in gs_asset.groupby("_anf_id"):
        statuses = grp["Umsetzungsgrad"].str.strip().tolist()
        worst    = min(statuses, key=lambda s: UZ_PRIORITY.get(s, 99))
        # Anforderungs-Typ: Basis / Standard / Erhöht
        typ_col = "Anforderungs-Typ"
        anf_typ = grp[typ_col].iloc[0].strip() if typ_col in grp.columns else ""
        rows.append({
            "anforderung_id":   anf_id,
            "anforderungs_typ": anf_typ,
            "baustein":         grp["Baustein"].iloc[0] if "Baustein" in grp.columns else "",
            "umsetzungsgrad":   worst,
            "all_statuses":     "; ".join(sorted(set(statuses))),
            "zielobjekte":      "; ".join(grp["Zielobjekt"].unique()),
        })
    return pd.DataFrame(rows)


# ─── CLASSIFY ────────────────────────────────────────────────────────────────

def classify_status(umsetzungsgrad: str, modalverb: str) -> str:
    """
    KANN        → Optional (regardless of Umsetzungsgrad)
    Ja          → Complied
    Entbehrlich → Exempted (review needed — GS exemption doesn't carry over to GS++)
    else        → Non-Complied
    """
    if modalverb == "KANN":
        return "🔵 Optional"
    if umsetzungsgrad == "Ja":
        return "✅ Complied"
    if umsetzungsgrad == "Entbehrlich":
        return "⚠️ Exempted"
    return "🔴 Non-Complied"


# ─── CORE ANALYSIS ───────────────────────────────────────────────────────────

def compute_asset_gspp_status(asset_reqs: pd.DataFrame, mapping_df: pd.DataFrame) -> pd.DataFrame:
    req_lookup     = dict(zip(asset_reqs["anforderung_id"], asset_reqs["umsetzungsgrad"]))
    req_typ_lookup = dict(zip(asset_reqs["anforderung_id"],
                              asset_reqs.get("anforderungs_typ", pd.Series("", index=asset_reqs.index))))

    rows = []
    for _, ctrl in mapping_df.iterrows():
        matched = {gs_id: req_lookup[gs_id]
                   for gs_id in ctrl["gs_ids_list"]
                   if gs_id in req_lookup}
        if not matched:
            continue

        worst_uz  = min(matched.values(), key=lambda s: UZ_PRIORITY.get(s, 99))
        mv        = ctrl.get("modalverb", "UNKNOWN")
        level     = ctrl.get("gspp_level", "normal-SdT")
        status    = classify_status(worst_uz, mv)
        priority  = compute_priority(status, mv, level)

        # Get Anforderungs-Typ for matched GS requirements
        matched_typs = "; ".join(
            req_typ_lookup.get(gs_id, "")
            for gs_id in matched.keys()
            if req_typ_lookup.get(gs_id,"")
        )

        rows.append({
            "priority":             priority,
            "gspp_control_id":      ctrl["gspp_id"],
            "gspp_control_title":   ctrl["gspp_title"],
            "gspp_practice":        ctrl["gspp_practice"],
            "gspp_level":           level,
            "modalverb":            mv,
            "asset_status":         status,
            "worst_umsetzungsgrad": worst_uz,
            "matched_gs_anforderungs_typ": matched_typs,
            "coverage":             ctrl.get("coverage",""),
            "confidence":           ctrl.get("confidence",""),
            "matched_gs_count":     len(matched),
            "matched_gs_ids":       "; ".join(matched.keys()),
            "matched_gs_statuses":  "; ".join(f"{k}[{v}]" for k, v in matched.items()),
            "rationale":            ctrl.get("rationale","")[:300] if ctrl.get("rationale","") not in ("N/A","") else "",
            "gap_notes":            ctrl.get("gap_notes","")[:300]  if ctrl.get("gap_notes","")  not in ("N/A","") else "",
            "model":                ctrl.get("model",""),
        })

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values("priority").reset_index(drop=True)


def compute_coverage(asset_reqs: pd.DataFrame, mapping_df: pd.DataFrame = None) -> dict:
    """
    GS++ coverage per Anforderungs-Typ (Basis/Standard/Erhöht).

    Only counts GS Anforderungen referenced in final_mapping.csv —
    i.e. only what GS++ actually requires from this asset.
    Ja = covered, everything else = not covered.

    Returns:
      {
        "Basis":    {"complied": 5, "non_complied": 2, "total": 7, "rate_%": 71.4},
        "Standard": {...},
        "Erhöht":   {...},
      }
    """
    if asset_reqs.empty or "anforderungs_typ" not in asset_reqs.columns:
        return {}

    # Restrict to only GS ids referenced by GS++ mapping
    if mapping_df is not None and not mapping_df.empty and "gs_ids_list" in mapping_df.columns:
        referenced = set(
            gs_id
            for gs_ids in mapping_df["gs_ids_list"]
            for gs_id in gs_ids
        )
        reqs = asset_reqs[asset_reqs["anforderung_id"].isin(referenced)].copy()
    else:
        reqs = asset_reqs.copy()

    if reqs.empty:
        return {}

    result = {}
    for level in ("Basis", "Standard", "Erhöht"):
        in_level = reqs[reqs["anforderungs_typ"].str.strip() == level]
        # Exclude Entbehrlich from coverage (counted separately as Exempted)
        in_level = in_level[in_level["umsetzungsgrad"] != "Entbehrlich"]
        complied     = int((in_level["umsetzungsgrad"] == "Ja").sum())
        non_complied = int((in_level["umsetzungsgrad"] != "Ja").sum())
        total        = complied + non_complied
        result[level] = {
            "complied":     complied,
            "non_complied": non_complied,
            "total":        total,
            "rate_%":       float(round(complied / total * 100, 1)) if total else 0.0,
        }
    return result


def summarize(gspp_pf: pd.DataFrame, asset_name: str, asset_reqs: pd.DataFrame = None, mapping_df: pd.DataFrame = None) -> dict:
    if gspp_pf.empty:
        return {"asset": asset_name, "total_controls": 0, "complied_rate_%": 0,
                "✅ Complied": 0, "🔴 Non-Complied": 0, "⚠️ Exempted": 0,
                "Basis_complied": 0, "Basis_total": 0, "Basis_rate_%": 0.0,
                "Standard_complied": 0, "Standard_total": 0, "Standard_rate_%": 0.0,
                "Erhöht_complied": 0, "Erhöht_total": 0, "Erhöht_rate_%": 0.0,
                "🔵 Optional": 0}
    counts = gspp_pf["asset_status"].value_counts().to_dict()
    total  = len(gspp_pf)
    result = {
        "asset":           asset_name,
        "total_controls":  total,
        "✅ Complied":     counts.get("✅ Complied", 0),
        "🔴 Non-Complied": counts.get("🔴 Non-Complied", 0),
        "complied_rate_%": float(round(counts.get("✅ Complied", 0) / total * 100, 1)),
        "⚠️ Exempted":    counts.get("⚠️ Exempted", 0),
    }
    if asset_reqs is not None:
        cov = compute_coverage(asset_reqs, mapping_df)
        for level in ("Basis", "Standard", "Erhöht"):
            result[f"{level}_complied"] = cov.get(level, {}).get("complied", 0)
            result[f"{level}_total"]    = cov.get(level, {}).get("total", 0)
            result[f"{level}_rate_%"]   = float(cov.get(level, {}).get("rate_%", 0.0))
    # Optional goes last
    result["🔵 Optional"] = counts.get("🔵 Optional", 0)
    return result


# ─── EXCEL ───────────────────────────────────────────────────────────────────

def _hdr(ws, cols, widths):
    fill = PatternFill("solid", start_color="1F3864")
    font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    if cols:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

def _write_df(ws, df, widths, status_col=None):
    if df is None or df.empty or len(df.columns) == 0:
        ws.cell(row=1, column=1, value="No data"); return
    _hdr(ws, list(df.columns), widths)
    for r, row in enumerate(df.itertuples(index=False), 2):
        sv = row[list(df.columns).index(status_col)] if status_col and status_col in df.columns else ""
        bg, fg = STATUS_STYLES.get(sv, ("FFFFFF","000000"))
        fl = PatternFill("solid", start_color=bg)
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fl
            cell.font = Font(name="Arial", size=9, color=fg)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

def _add_legend(wb):
    # Insert legend as first sheet using index=0
    ws = wb.create_sheet("📖 Legend", index=0)
    cols   = ["Status","Meaning","Modalverb","GS Umsetzungsgrad","Recommended Action"]
    widths = [22,45,20,20,42]
    _hdr(ws, cols, widths)
    for r, row in enumerate(LEGEND_ROWS, 2):
        bg, fg = STATUS_STYLES.get(row[0], ("FFFFFF","000000"))
        fl = PatternFill("solid", start_color=bg)
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fl
            cell.font = Font(name="Arial", size=10, bold=(c==1), color=fg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 30


# ─── COVERAGE SHEET ──────────────────────────────────────────────────────────

def _write_coverage_sheet(ws, cov: dict):
    """Write coverage breakdown by Basis/Standard/Erhöht to a worksheet."""
    LEVELS = ["Basis", "Standard", "Erhöht"]
    fill_hdr = PatternFill("solid", start_color="1F3864")
    font_hdr = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    font_reg = Font(name="Arial", size=9)

    # Header
    headers = ["Scope"] + [f"{l} Complied" for l in LEVELS] +               [f"{l} Total" for l in LEVELS] + [f"{l} Rate %" for l in LEVELS]
    widths  = [30] + [14]*9
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_hdr; c.font = font_hdr
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    def write_row(ws, r, scope, data):
        row_vals = [scope]
        for l in LEVELS: row_vals.append(data.get(l,{}).get("complied",0))
        for l in LEVELS: row_vals.append(data.get(l,{}).get("total",0))
        for l in LEVELS:
            rate = data.get(l,{}).get("rate_%",0)
            row_vals.append(f"{rate}%")
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = font_reg
            cell.alignment = Alignment(vertical="top")

    # Overall row
    write_row(ws, 2, "Overall", cov)

    # Per practice rows
    r = 3
    for practice, pdata in sorted(cov.get("by_practice", {}).items()):
        write_row(ws, r, f"  {practice}", pdata)
        r += 1


# ─── SINGLE ASSET ────────────────────────────────────────────────────────────

PF_COLS   = ["priority","gspp_control_id","gspp_control_title","gspp_practice",
             "gspp_level","modalverb","asset_status","worst_umsetzungsgrad",
             "matched_gs_anforderungs_typ","coverage",
             "confidence","matched_gs_count","matched_gs_ids","matched_gs_statuses",
             "rationale","gap_notes","model"]
PF_WIDTHS = [8,16,42,12,12,10,20,16,20,12,10,8,55,55,70,70,20]

def run(asset_name: str, gs_check: pd.DataFrame, mapping_df: pd.DataFrame,
        zo_df: pd.DataFrame = None, cache: dict = None):

    gs_raw = get_asset_requirements(asset_name, gs_check)
    if gs_raw.empty:
        print(f"  No GS requirements found for '{asset_name}'")
        return None, None

    asset_reqs = aggregate_requirements(gs_raw)
    status_dist = asset_reqs["umsetzungsgrad"].value_counts().to_dict()
    print(f"  GS reqs: {len(asset_reqs)}  {status_dist}")

    # Zielobjekt matching
    if zo_df is not None and cache is not None:
        zo_matches = match_zielobjekt(asset_name, zo_df, cache)
        top = zo_matches[0] if zo_matches else {}
        print(f"  Zielobjekt: {top.get('zielobjekt','?')} "
              f"({top.get('typ','')}) score={top.get('score',0):.3f}")

    gspp_pf = compute_asset_gspp_status(asset_reqs, mapping_df)
    if gspp_pf.empty:
        print(f"  No GS++ controls matched.")
        return asset_reqs, gspp_pf

    s = summarize(gspp_pf, asset_name, asset_reqs, mapping_df)
    print(f"  GS++ controls: {s['total_controls']}  "
          f"✅{s['✅ Complied']} 🔴{s['🔴 Non-Complied']} 🔵{s['🔵 Optional']}  "
          f"Complied: {s['complied_rate_%']}%")

    gaps = gspp_pf[gspp_pf["asset_status"] == "🔴 Non-Complied"]
    if not gaps.empty:
        print(f"  Non-Complied (priority order):")
        for _, row in gaps.iterrows():
            print(f"    [P{row['priority']}] {row['gspp_control_id']:<16} "
                  f"[{row['modalverb']}/{row['gspp_level']}]  "
                  f"{row['gspp_control_title'][:40]}")

    return asset_reqs, gspp_pf


def build_detailed_matches(asset_name: str, gs_reqs: pd.DataFrame,
                           gspp_pf: pd.DataFrame, mapping_df: pd.DataFrame) -> pd.DataFrame:
    """
    Expand matched_gs_ids into one row per GS++ control × GS Anforderung pair.
    """
    req_lookup     = dict(zip(gs_reqs["anforderung_id"], gs_reqs["umsetzungsgrad"]))
    req_typ_lookup = dict(zip(gs_reqs["anforderung_id"],
                              gs_reqs.get("anforderungs_typ", pd.Series("", index=gs_reqs.index))))

    # Build mapping lookup: gspp_id → gs_ids_list
    map_lookup = {}
    for _, ctrl in mapping_df.iterrows():
        map_lookup[ctrl["gspp_id"]] = ctrl["gs_ids_list"]

    rows = []
    for _, ctrl in gspp_pf.iterrows():
        gspp_id  = ctrl["gspp_control_id"]
        gs_ids   = map_lookup.get(gspp_id, [])
        matched  = [g for g in gs_ids if g in req_lookup]

        for gs_id in matched:
            uz  = req_lookup.get(gs_id, "")
            typ = req_typ_lookup.get(gs_id, "")
            rows.append({
                "asset":              asset_name,
                "gspp_control_id":    gspp_id,
                "gspp_control_title": ctrl["gspp_control_title"],
                "gspp_practice":      ctrl["gspp_practice"],
                "gspp_level":         ctrl["gspp_level"],
                "modalverb":          ctrl["modalverb"],
                "asset_status":       ctrl["asset_status"],
                "priority":           ctrl["priority"],
                "gs_anforderung_id":  gs_id,
                "anforderungs_typ":   typ,
                "gs_umsetzungsgrad":  uz,
                "gs_complied":        "✅" if uz == "Ja" else ("⚠️" if uz == "Entbehrlich" else "🔴"),
            })

    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    return df.sort_values(["asset_status","gspp_control_id"]).reset_index(drop=True)


DET_COLS   = ["asset","gspp_control_id","gspp_control_title","gspp_practice",
              "gspp_level","modalverb","asset_status","priority",
              "gs_anforderung_id","anforderungs_typ","gs_umsetzungsgrad","gs_complied"]
DET_WIDTHS = [35,16,42,12,12,10,20,8,18,14,14,8]


def write_single_xlsx(asset_name, gs_reqs, gspp_pf, mapping_df, path):
    wb = Workbook()
    del wb[wb.sheetnames[0]]  # remove default empty sheet
    _add_legend(wb)

    ws1 = wb.active; ws1.title = "GS Requirements"
    cols1 = ["anforderung_id","anforderungs_typ","baustein","umsetzungsgrad","all_statuses","zielobjekte"]
    _write_df(ws1, gs_reqs[[c for c in cols1 if c in gs_reqs.columns]],
              [18,14,14,14,20,40], "umsetzungsgrad")

    ws2 = wb.create_sheet("GS++ Pass-Fail")
    _write_df(ws2, gspp_pf[[c for c in PF_COLS if c in gspp_pf.columns]],
              PF_WIDTHS, "asset_status")

    # Detailed Matches sheet
    det = build_detailed_matches(asset_name, gs_reqs, gspp_pf, mapping_df)
    if not det.empty:
        ws3 = wb.create_sheet("Detailed Matches")
        _write_df(ws3, det[[c for c in DET_COLS if c in det.columns]],
                  DET_WIDTHS, "asset_status")

    # Coverage sheet
    cov = compute_coverage(gs_reqs, mapping_df)
    if cov:
        ws4 = wb.create_sheet("Coverage")
        _write_coverage_sheet(ws4, cov)

    wb.save(str(path))
    print(f"  Excel: {path}")


# ─── BATCH RUN ────────────────────────────────────────────────────────────────

def run_all(gs_check: pd.DataFrame, mapping_df: pd.DataFrame,
            zo_df: pd.DataFrame = None, cache: dict = None):

    all_assets = gs_check["Zielobjekt"].value_counts().index.tolist()
    assets     = [a for a in all_assets if a not in GENERAL_ASSETS]
    general    = [a for a in all_assets if a in GENERAL_ASSETS]
    print(f"\nBATCH: {len(assets)} assets + {len(general)} general")

    all_pf, general_pf, summaries, all_det_list = [], [], [], []

    for asset in general:
        print(f"\n→ [GENERAL] {asset}")
        _, pf = run(asset, gs_check, mapping_df, zo_df, cache)
        if pf is not None and not pf.empty:
            pf["asset"] = asset; general_pf.append(pf)

    for asset in assets:
        print(f"\n→ {asset}")
        gs_reqs, pf = run(asset, gs_check, mapping_df, zo_df, cache)
        if pf is None: continue
        summaries.append(summarize(pf, asset, gs_reqs, mapping_df))
        if not pf.empty:
            pf["asset"] = asset
            all_pf.append(pf)
            det = build_detailed_matches(asset, gs_reqs, pf, mapping_df)
            if not det.empty:
                all_det_list.append(det)

    all_pf_df     = pd.concat(all_pf,     ignore_index=True) if all_pf     else pd.DataFrame()
    general_pf_df = pd.concat(general_pf, ignore_index=True) if general_pf else pd.DataFrame()
    summary_df    = pd.DataFrame(summaries)

    print(f"\n{'='*60}\nBATCH SUMMARY\n{'='*60}")
    if not summary_df.empty:
        print(f"  {'Asset':<45} {'Complied%':>9}  {'Gaps':>5}")
        for _, row in summary_df.sort_values("complied_rate_%").iterrows():
            print(f"  {row['asset']:<45} {row['complied_rate_%']:>8.1f}%  {row['🔴 Non-Complied']:>5}")

    BATCH_PF_COLS = ["asset"] + PF_COLS
    BATCH_WIDTHS  = [35] + PF_WIDTHS

    wb = Workbook()
    del wb[wb.sheetnames[0]]  # remove default empty sheet
    _add_legend(wb)

    ws1 = wb.create_sheet("Summary")
    if not summary_df.empty:
        n_cols = len(summary_df.columns)
        widths = [45] + [12] * (n_cols - 1)
        _write_df(ws1, summary_df, widths)

    ws2 = wb.create_sheet("GS++ Pass-Fail (All)")
    if not all_pf_df.empty:
        _write_df(ws2, all_pf_df[[c for c in BATCH_PF_COLS if c in all_pf_df.columns]],
                  BATCH_WIDTHS, "asset_status")

    # Detailed Matches (All)
    all_det = pd.concat(all_det_list, ignore_index=True) if all_det_list else pd.DataFrame()
    ws_det = wb.create_sheet("Detailed Matches (All)")
    if not all_det.empty:
        _write_df(ws_det, all_det[[c for c in DET_COLS if c in all_det.columns]],
                  DET_WIDTHS, "asset_status")

    ws3 = wb.create_sheet("Informationsverbund")
    if not general_pf_df.empty:
        _write_df(ws3, general_pf_df[[c for c in BATCH_PF_COLS if c in general_pf_df.columns]],
                  BATCH_WIDTHS, "asset_status")

    xlsx_out = OUT / "asset_gap_ALL.xlsx"
    wb.save(str(xlsx_out))
    print(f"\nExcel: {xlsx_out}")

    json_out = OUT / "asset_gap_ALL.json"
    with open(str(json_out),"w",encoding="utf-8") as f:
        json.dump(summaries, f, ensure_ascii=False, indent=2)
    print(f"JSON:  {json_out}")


# ─── JSON INPUT ───────────────────────────────────────────────────────────────

def load_from_json(json_path: str) -> pd.DataFrame:
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    rows = []
    if isinstance(data, dict) and "anforderungen" in data:
        for item in data["anforderungen"]:
            rows.append({"Zielobjekt":     item.get("zielobjekt",""),
                         "Anforderung":    item.get("anforderung_id",""),
                         "Umsetzungsgrad": item.get("umsetzungsgrad","Nicht geprüft"),
                         "Baustein":       item.get("baustein_id","")})
    elif isinstance(data, dict) and "assets" in data:
        for asset, anf_list in data["assets"].items():
            for item in anf_list:
                rows.append({"Zielobjekt":     asset,
                             "Anforderung":    item.get("anforderung_id",""),
                             "Umsetzungsgrad": item.get("umsetzungsgrad",item.get("status","Nicht geprüft")),
                             "Baustein":       item.get("baustein","")})
    elif isinstance(data, list):
        for item in data:
            rows.append({"Zielobjekt":     item.get("zielobjekt",item.get("asset","")),
                         "Anforderung":    item.get("anforderung_id",item.get("id","")),
                         "Umsetzungsgrad": item.get("umsetzungsgrad",item.get("status","Nicht geprüft")),
                         "Baustein":       item.get("baustein","")})
    df = pd.DataFrame(rows)
    print(f"  JSON: {len(df)} rows | {df['Zielobjekt'].nunique()} assets")
    return df


# ─── CLI ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Asset-Based GS to GS++ Gap Analysis")
    parser.add_argument("--asset",    "-a", default=None)
    parser.add_argument("--all",      action="store_true")
    parser.add_argument("--json-file","-j", default=None)
    parser.add_argument("--mapping",  "-m", default=str(FINAL_MAPPING))
    args = parser.parse_args()

    # Load GS data
    gs_check = load_from_json(args.json_file) if args.json_file else load_gs_check_csv()
    print(f"GS data: {len(gs_check)} rows | {gs_check['Zielobjekt'].nunique()} assets")

    # Load mapping + modalverb
    print("Loading modalverb map from catalog...")
    mv_map = load_modalverb_map()
    mapping_df = load_final_mapping(Path(args.mapping), mv_map)

    # Load Zielobjekt data + cache for embedding
    zo_df = pd.read_csv(str(ZIELOBJEKT_CSV), dtype=str).fillna("") if ZIELOBJEKT_CSV.exists() else None
    cache = load_cache() if zo_df is not None else None
    if zo_df is not None:
        print(f"Zielobjekt categories: {len(zo_df)}")

    if args.all:
        run_all(gs_check, mapping_df, zo_df, cache)
    elif args.asset:
        print(f"\n{'='*60}\nAsset: '{args.asset}'\n{'='*60}")
        gs_reqs, gspp_pf = run(args.asset, gs_check, mapping_df, zo_df, cache)
        if gspp_pf is not None and not gspp_pf.empty:
            safe = args.asset.lower().replace(" ","_").replace("/","_")
            write_single_xlsx(args.asset, gs_reqs, gspp_pf, mapping_df, OUT/f"asset_gap_{safe}.xlsx")
    else:
        print("Available assets:")
        for zo, cnt in gs_check["Zielobjekt"].value_counts().items():
            tag = " [GENERAL]" if zo in GENERAL_ASSETS else ""
            print(f"  {zo} ({cnt}){tag}")
        print("\nUsage:")
        print("  python3 src/asset_gap.py --asset 'Virtualisierungsserver'")
        print("  python3 src/asset_gap.py --all")